# modules/scanner_handicap.py
# =====================================================
# SCANNER B - Value bet sugli handicap games
# Calibrazione empirica da 2500 partite ATP 2025
# =====================================================

import os
import re
import sys
import time
import threading
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from math import erf, sqrt
from datetime import datetime

# Forza UTF-8 su Windows per le emoji nelle print
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from modules.elo_tennisabstract import carica_elo_aggiornato, trova_giocatore_ta, verifica_disponibilita
from modules.odds_apiio import get_partite_con_quote_oggi, get_quote_handicap_evento, get_quote_handicap_sets_evento
from modules.forma_recente import carica_partite_recenti, calcola_forma, aggiusta_elo_per_forma, calcola_h2h, calcola_fatica
from modules.data_2025 import scarica_e_salva_2026
from modules.scanner import normalizza_superficie, superficie_da_torneo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from config import ODDS_API_IO_KEY, EV_MAX, EV_MINIMO, ODDS_MIN_VALUE, ODDS_MAX_VALUE
except Exception:
    ODDS_API_IO_KEY = ""
    EV_MAX = None
    EV_MINIMO = 0.09
    ODDS_MIN_VALUE = 1.80
    ODDS_MAX_VALUE = 4.00

SIGMA = 4.8  # deviazione standard calibrata su 2500 partite ATP 2025

# Punti di ancoraggio per l'interpolazione lineare:
# (diff_elo, margine_atteso_games) dal punto di vista del favorito
# Derivati dalla tabella empirica: midpoint di ogni banda + punto 0
CALIBRAZIONE = [
    (0,   0.0),
    (25,  0.4),
    (75,  1.1),
    (125, 2.1),
    (175, 2.6),
    (250, 4.2),
]

# Calibrazione empirica set: (diff_elo_midpoint, P(favorito vince 2-0))
CALIBRAZIONE_SETS = [
    (0,   0.35),
    (25,  0.35),
    (75,  0.42),
    (125, 0.50),
    (175, 0.58),
    (250, 0.65),
]


# ---------------------------------------------------------------------------
# Utility matematiche
# ---------------------------------------------------------------------------

def _norm_cdf(x: float) -> float:
    return (1.0 + erf(x / sqrt(2.0))) / 2.0


def margine_atteso_da_diff(diff_elo: float) -> float:
    """
    Interpolazione lineare sulla tabella di calibrazione.
    Restituisce il margine games atteso dal punto di vista del favorito.
    """
    if diff_elo <= CALIBRAZIONE[0][0]:
        return CALIBRAZIONE[0][1]
    if diff_elo >= CALIBRAZIONE[-1][0]:
        return CALIBRAZIONE[-1][1]
    for i in range(len(CALIBRAZIONE) - 1):
        x0, y0 = CALIBRAZIONE[i]
        x1, y1 = CALIBRAZIONE[i + 1]
        if x0 <= diff_elo <= x1:
            t = (diff_elo - x0) / (x1 - x0)
            return y0 + t * (y1 - y0)
    return CALIBRAZIONE[-1][1]


def prob_2_0_da_diff(diff_elo: float) -> float:
    """Interpolazione: P(favorito Elo vince 2-0 nei set) dato diff Elo."""
    if diff_elo <= CALIBRAZIONE_SETS[0][0]:
        return CALIBRAZIONE_SETS[0][1]
    if diff_elo >= CALIBRAZIONE_SETS[-1][0]:
        return CALIBRAZIONE_SETS[-1][1]
    for i in range(len(CALIBRAZIONE_SETS) - 1):
        x0, y0 = CALIBRAZIONE_SETS[i]
        x1, y1 = CALIBRAZIONE_SETS[i + 1]
        if x0 <= diff_elo <= x1:
            t = (diff_elo - x0) / (x1 - x0)
            return y0 + t * (y1 - y0)
    return CALIBRAZIONE_SETS[-1][1]


def prob_copre_sets(h: float, is_fav: bool, p2_0_fav: float):
    """
    P(cover) per handicap sets ±1.5.
    h: handicap di questo giocatore (-1.5 o +1.5)
    is_fav: True se questo giocatore è il favorito Elo
    p2_0_fav: P(favorito Elo vince 2-0) dalla tabella di calibrazione
    Restituisce None per linee non standard o combinazioni incoerenti.
    """
    if abs(abs(h) - 1.5) > 0.01:
        return None
    if h < 0:  # questo giocatore deve vincere 2-0
        return p2_0_fav if is_fav else None
    else:  # copre se non perde 0-2 (basta vincere almeno un set)
        return (1.0 - p2_0_fav) if not is_fav else None


def prob_copre(handicap: float, margine_atteso: float, sigma: float = SIGMA) -> float:
    """
    P(margine_reale > -handicap)
    margine_atteso: dal punto di vista del giocatore valutato
                   (positivo se è il favorito, negativo se è l'underdog)
    handicap: il valore di handicap assegnato a quel giocatore
              (es. -3.5 per il favorito, +3.5 per l'underdog)
    """
    soglia = -handicap
    z = (soglia - margine_atteso) / sigma
    return 1.0 - _norm_cdf(z)


# ---------------------------------------------------------------------------
# Core: analisi handicap per lista di partite
# ---------------------------------------------------------------------------

def analizza_handicap(partite, ratings_ta, soglia_ev,
                      partite_recenti=None, nomi_recenti_dict=None):
    value_bets = []
    senza_quote = []
    non_trovati = []

    # ── Rate limiter: max 10 richieste al secondo ──────────────────────
    _rl_lock = threading.Lock()
    _rl_slots = []

    def _rate_limit():
        with _rl_lock:
            while True:
                now = time.time()
                _rl_slots[:] = [t for t in _rl_slots if now - t < 1.0]
                if len(_rl_slots) < 10:
                    break
                time.sleep(1.0 - (now - _rl_slots[0]) + 0.001)
            _rl_slots.append(time.time())

    def _fetch_quotes(event_id):
        _rate_limit()
        linee = get_quote_handicap_evento(event_id)
        _rate_limit()
        linee_sets = get_quote_handicap_sets_evento(event_id)
        return linee, linee_sets

    # ── FASE 1: calcolo Elo (sequenziale, in-memory) ───────────────────
    partite_valide = []
    for p in partite:
        sup_raw = p.get('superficie', '')
        sup = normalizza_superficie(sup_raw)
        if sup == 'hard' and not sup_raw:
            sup = superficie_da_torneo(p.get('torneo', ''))

        n1, r1 = trova_giocatore_ta(p['p1'], ratings_ta)
        n2, r2 = trova_giocatore_ta(p['p2'], ratings_ta)
        if not r1 or not r2:
            mancante = p['p1'] if not r1 else p['p2']
            non_trovati.append(f"{p['p1']} vs {p['p2']} (manca: {mancante})")
            continue

        e1 = r1.get(sup, r1['elo'])
        e2 = r2.get(sup, r2['elo'])
        elo_usato = f"TA-{sup}"

        n1_fr = n2_fr = None
        if partite_recenti and nomi_recenti_dict:
            n1_fr, _ = trova_giocatore_ta(n1, nomi_recenti_dict)
            n2_fr, _ = trova_giocatore_ta(n2, nomi_recenti_dict)
            forma1 = calcola_forma(n1_fr, partite_recenti, superficie=sup) if n1_fr else 0.0
            forma2 = calcola_forma(n2_fr, partite_recenti, superficie=sup) if n2_fr else 0.0
            e1 = aggiusta_elo_per_forma(e1, forma1)
            e2 = aggiusta_elo_per_forma(e2, forma2)
            elo_usato = f"TA-{sup}+forma"

        # H2H — aggiusta Elo di max ±50 punti (ridotto da 75; pochi dati su Challenger)
        h2h = calcola_h2h(n1, n2, partite_recenti)
        e1 += h2h * 50
        e2 -= h2h * 50
        if h2h != 0:
            elo_usato += f"+h2h({h2h:+.2f})"

        # Fatica: max -100 punti Elo; si attiva con 2+ partite negli ultimi 2 giorni
        if partite_recenti:
            fatica1 = calcola_fatica(n1_fr if n1_fr else n1, partite_recenti)
            fatica2 = calcola_fatica(n2_fr if n2_fr else n2, partite_recenti)
            e1 += fatica1 * 100
            e2 += fatica2 * 100
            if fatica1 != 0 or fatica2 != 0:
                elo_usato += f"+fat({fatica1:.1f}/{fatica2:.1f})"

        diff_elo = abs(e1 - e2)
        margine_fav = margine_atteso_da_diff(diff_elo)
        margine_home = margine_fav if e1 >= e2 else -margine_fav

        partite_valide.append({
            'p': p, 'e1': e1, 'e2': e2,
            'sup': sup, 'elo_usato': elo_usato,
            'diff_elo': diff_elo, 'margine_fav': margine_fav, 'margine_home': margine_home,
        })

    # ── FASE 2: fetch quote in parallelo (2 call/partita, rate-limited) ──
    quote_cache = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {
            executor.submit(_fetch_quotes, item['p']['id']): item['p']['id']
            for item in partite_valide
        }
        for future in as_completed(futures):
            eid = futures[future]
            try:
                quote_cache[eid] = future.result()
            except Exception:
                quote_cache[eid] = (None, None)

    # ── FASE 3: segnali, filtro range Pro, value bet ───────────────────
    for item in partite_valide:
        p = item['p']
        e1, e2 = item['e1'], item['e2']
        sup, elo_usato = item['sup'], item['elo_usato']
        diff_elo, margine_fav, margine_home = (
            item['diff_elo'], item['margine_fav'], item['margine_home']
        )

        linee, linee_sets = quote_cache.get(p['id'], (None, None))

        if not linee:
            senza_quote.append(f"{p['p1']} vs {p['p2']}")
        else:
            for linea in linee:
                h_home = linea['handicap']
                q_home = linea['quota_home']
                q_away = linea['quota_away']

                pc_home = prob_copre(h_home, margine_home)
                ev_home = pc_home * q_home - 1.0
                pc_away = prob_copre(-h_home, -margine_home)
                ev_away = pc_away * q_away - 1.0

                for player, opp, ev, quota, pc, h in [
                    (p['p1'], p['p2'], ev_home, q_home, pc_home, h_home),
                    (p['p2'], p['p1'], ev_away, q_away, pc_away, -h_home),
                ]:
                    if ev >= soglia_ev and ODDS_MIN_VALUE <= quota <= ODDS_MAX_VALUE:
                        value_bets.append({
                            "p1": player,
                            "p2": opp,
                            "torneo": p['torneo'],
                            "superficie": sup,
                            "handicap": h,
                            "quota_handicap": quota,
                            "prob_stimata": round(pc, 4),
                            "ev": round(ev, 4),
                            "elo_diff": round(diff_elo, 0),
                            "margine_atteso": round(margine_fav, 2),
                            "elo_usato": elo_usato,
                            "source": p.get('source', ''),
                            "data_partita": p.get('data_partita', ''),
                            "tipo": "Games",
                        })

        if linee_sets:
            home_is_fav = (e1 >= e2)
            p2_0 = prob_2_0_da_diff(diff_elo)
            for linea in linee_sets:
                h_home = linea['handicap']
                q_home = linea['quota_home']
                q_away = linea['quota_away']
                for player, opp, pc, quota, h in [
                    (p['p1'], p['p2'], prob_copre_sets(h_home, home_is_fav, p2_0), q_home, h_home),
                    (p['p2'], p['p1'], prob_copre_sets(-h_home, not home_is_fav, p2_0), q_away, -h_home),
                ]:
                    if pc is None:
                        continue
                    ev = pc * quota - 1.0
                    if ev >= soglia_ev and ODDS_MIN_VALUE <= quota <= ODDS_MAX_VALUE:
                        value_bets.append({
                            "p1": player,
                            "p2": opp,
                            "torneo": p['torneo'],
                            "superficie": sup,
                            "handicap": h,
                            "quota_handicap": quota,
                            "prob_stimata": round(pc, 4),
                            "ev": round(ev, 4),
                            "elo_diff": round(diff_elo, 0),
                            "margine_atteso": round(p2_0, 4),
                            "elo_usato": elo_usato,
                            "source": p.get('source', ''),
                            "data_partita": p.get('data_partita', ''),
                            "tipo": "Sets",
                        })

    return value_bets, senza_quote, non_trovati


# ---------------------------------------------------------------------------
# Salvataggio Excel — sheet "Handicap Bets"
# ---------------------------------------------------------------------------

def salva_handicap_excel(value_bets, filepath="data/value_bets_log.xlsx"):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    SHEET_NAME = "Handicap Bets"
    blu_scuro   = PatternFill("solid", fgColor="1E4E79")
    blu_chiaro  = PatternFill("solid", fgColor="DDEEFF")
    bianco      = PatternFill("solid", fgColor="FFFFFF")
    giallo      = PatternFill("solid", fgColor="FFEB9C")
    azzurro     = PatternFill("solid", fgColor="BDD7EE")
    bordo = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        wb.active.title = "Value Bets Log"

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        # Raccogli TUTTE le partite presenti (indipendentemente dalla data)
        # e rimuovi eventuali duplicati già nell'Excel
        partite_esistenti = set()
        righe_duplicate = []
        for row_idx in range(2, ws.max_row + 1):
            data_val = ws.cell(row_idx, 1).value
            if not data_val:
                continue
            data_str = data_val.strftime("%Y-%m-%d") if hasattr(data_val, 'strftime') else str(data_val).strip()
            p1_raw = str(ws.cell(row_idx, 3).value or '')
            p1_raw = re.sub(r'^✅\s+', '', p1_raw)
            p1_raw = re.sub(r'\s+[+-]\d+\.?\d*$', '', p1_raw).strip()
            p2_raw   = str(ws.cell(row_idx, 4).value or '').strip()
            try:
                h_raw = f"{float(ws.cell(row_idx, 6).value):.1f}"
            except (TypeError, ValueError):
                h_raw = str(ws.cell(row_idx, 6).value or '').strip()
            tipo_val = ws.cell(row_idx, 5).value or 'Games'
            key = (data_str, p1_raw, p2_raw, h_raw, tipo_val)
            if key in partite_esistenti:
                righe_duplicate.append(row_idx)
            else:
                partite_esistenti.add(key)

        for row_idx in reversed(righe_duplicate):
            ws.delete_rows(row_idx)
        if righe_duplicate:
            print(f"  🧹 {len(righe_duplicate)} righe duplicate rimosse dall'Excel")

        prossima_riga = ws.max_row + 1
    else:
        ws = wb.create_sheet(SHEET_NAME)
        partite_esistenti = set()
        headers = [
            "Data", "Ora", "Punta su (handicap)", "Avversario",
            "Tipo", "Handicap", "Quota", "Quota di chiusura", "Prob %", "EV %",
            "Torneo", "Superficie", "Elo Diff", "Margine Atteso",
            "Elo Usato", "Fonte",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = blu_scuro
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bordo
        ws.row_dimensions[1].height = 28
        prossima_riga = 2

    ora_now = datetime.now()
    aggiunte = saltate = 0

    for v in value_bets:
        data_p = v.get('data_partita', '').strip()
        key = (data_p, v['p1'].strip(), v['p2'].strip(), f"{float(v['handicap']):.1f}", v.get('tipo', 'Games'))
        if key in partite_esistenti:
            saltate += 1
            continue
        partite_esistenti.add(key)

        riga = prossima_riga + aggiunte
        fill = blu_chiaro if aggiunte % 2 == 0 else bianco

        valori = [
            v.get('data_partita', ora_now.strftime("%Y-%m-%d")),
            ora_now.strftime("%H:%M"),
            f"✅ {v['p1']}  {v['handicap']:+.1f}",
            v['p2'],
            v.get('tipo', 'Games'),
            v['handicap'],
            v['quota_handicap'],
            '',
            round(v['prob_stimata'] * 100, 1),
            round(v['ev'] * 100, 1),
            v['torneo'],
            v['superficie'].upper(),
            int(v['elo_diff']),
            v['margine_atteso'],
            v.get('elo_usato', ''),
            v.get('source', ''),
        ]

        for col, val in enumerate(valori, 1):
            cell = ws.cell(row=riga, column=col, value=val)
            cell.border = bordo
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 3:
                cell.fill = azzurro
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == 7:
                cell.fill = fill
                cell.font = Font(bold=True, size=12)
            elif col == 10 and isinstance(val, (int, float)):
                if val >= 30:
                    cell.fill = PatternFill("solid", fgColor="FF0000")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val >= 15:
                    cell.fill = PatternFill("solid", fgColor="FFC000")
                    cell.font = Font(bold=True)
                elif val >= 5:
                    cell.fill = giallo
                else:
                    cell.fill = fill
            else:
                cell.fill = fill
        ws.row_dimensions[riga].height = 22
        aggiunte += 1

    larghezze = [12, 8, 32, 22, 8, 10, 8, 18, 8, 8, 30, 10, 10, 14, 20, 14]
    for col, width in enumerate(larghezze, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"\n💾 {aggiunte} handicap bet salvate | {saltate} duplicati saltati → {filepath} [{SHEET_NAME}]")


# ---------------------------------------------------------------------------
# Entry point principale
# ---------------------------------------------------------------------------

def scansiona_handicap(soglia_ev=EV_MINIMO, ev_max=None):
    print(f"{'='*65}")
    print(f"  🎯 SCANNER HANDICAP - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*65}\n")

    # Safety check
    print("🔒 Safety check fonti dati...")
    ok_ta, msg_ta = verifica_disponibilita()
    print(f"  Tennis Abstract : {'✅' if ok_ta else '❌'} {msg_ta}")

    ok_odds = False
    if ODDS_API_IO_KEY:
        try:
            r = requests.get(
                "https://api.odds-api.io/v3/sports",
                params={"apiKey": ODDS_API_IO_KEY}, timeout=5
            )
            ok_odds = r.status_code == 200
            print(f"  Odds-API.io     : {'✅' if ok_odds else '❌'} Status {r.status_code}")
        except Exception as e:
            print(f"  Odds-API.io     : ❌ {e}")

    if not ok_ta:
        print("\n🚨 BLOCCO: Tennis Abstract non disponibile.")
        return []
    if not ok_odds:
        print("\n🚨 BLOCCO: Odds-API.io non disponibile.")
        return []

    print("\n✅ Fonti disponibili — procedo\n")

    # Carica Elo
    print("⚙️  Caricamento Elo Tennis Abstract...")
    ratings_ta = carica_elo_aggiornato()
    if not ratings_ta:
        print("🚨 BLOCCO: Nessun dato Elo — scanner fermato.")
        return []
    print(f"✅ {len(ratings_ta)} giocatori caricati\n")

    # Dati 2026
    csv_2026 = 'data/storico/atp_2026_tml.csv'
    deve_scaricare = (
        not os.path.exists(csv_2026)
        or (datetime.now() - datetime.fromtimestamp(
            os.path.getmtime(csv_2026))).total_seconds() > 86400
    )
    if deve_scaricare:
        print("📥 Aggiornamento dati 2026...")
        try:
            scarica_e_salva_2026()
        except Exception as e:
            print(f"  ⚠️  Download 2026 fallito: {e}")
    else:
        print("📦 Dati 2026 già aggiornati")

    # Risultati recenti
    csv_recenti = 'data/storico/risultati_recenti.csv'
    deve_aggiornare = (
        not os.path.exists(csv_recenti)
        or (datetime.now() - datetime.fromtimestamp(
            os.path.getmtime(csv_recenti))).total_seconds() > 86400
    )
    if deve_aggiornare:
        try:
            from modules.data_2025 import scarica_risultati_recenti_tennisexplorer
            scarica_risultati_recenti_tennisexplorer(giorni=60)
        except Exception as e:
            print(f"  ⚠️  Download risultati recenti fallito: {e}")

    print("📈 Caricamento forma recente (ultimi 30 giorni)...")
    partite_recenti = carica_partite_recenti(giorni=30)
    nomi_recenti_dict = {nome: {'elo': 0} for nome in partite_recenti}
    print(f"✅ Forma disponibile per {len(partite_recenti)} giocatori\n")

    # Recupera partite
    print("📅 Recupero partite da Odds-API.io...")
    partite_apiio = get_partite_con_quote_oggi()
    TORNEI_VALIDI  = ['ATP -', 'WTA -', 'Challenger -']
    TORNEI_ESCLUDI = ['125K', '125k', 'UTR', 'ITF', 'WTA 125']
    partite_apiio = [
        p for p in partite_apiio
        if any(x in p['torneo'] for x in TORNEI_VALIDI)
        and not any(x in p['torneo'] for x in TORNEI_ESCLUDI)
    ]
    print(f"✅ {len(partite_apiio)} partite filtrate\n")

    # Analisi handicap
    print("🔍 Analisi quote handicap...\n")
    value_bets, senza_quote, non_trovati = analizza_handicap(
        partite_apiio, ratings_ta, soglia_ev,
        partite_recenti, nomi_recenti_dict,
    )

    # Filtro EV massimo
    if ev_max is not None:
        prima = len(value_bets)
        value_bets = [v for v in value_bets if v['ev'] <= ev_max]
        escluse = prima - len(value_bets)
        if escluse:
            print(f"⚠️  {escluse} value bet escluse per EV > {ev_max*100:.0f}%\n")

    # Deduplicazione (stessa coppia + stesso handicap)
    seen: set = set()
    dedup = []
    for v in sorted(value_bets, key=lambda x: x['ev'], reverse=True):
        key = (v['p1'], v['p2'], v['handicap'], v.get('tipo', 'Games'))
        if key not in seen:
            seen.add(key)
            dedup.append(v)
    value_bets = dedup

    # Output
    print(f"{'='*65}")
    print(f"  RISULTATI SCANNER HANDICAP")
    print(f"{'='*65}")
    print(f"  Partite analizzate   : {len(partite_apiio)}")
    print(f"  Senza quote handicap : {len(senza_quote)}")
    print(f"  Non trovati in Elo   : {len(non_trovati)}")
    print(f"  VALUE BET ✅         : {len(value_bets)}")
    print(f"{'='*65}\n")

    if value_bets:
        print("🎯 VALUE BET HANDICAP TROVATE:\n")
        for v in value_bets:
            tipo = v.get('tipo', 'Games')
            print(f"  ✅ {v['p1']} ({v['handicap']:+.1f} {tipo.lower()})  vs  {v['p2']}")
            print(f"     Torneo       : {v['torneo']}")
            print(f"     Superficie   : {v['superficie'].upper()} | Elo diff: {int(v['elo_diff'])}")
            if tipo == 'Sets':
                print(f"     P(2-0 fav)   : {v['margine_atteso']*100:.1f}%")
            else:
                print(f"     Margine fav  : {v['margine_atteso']:+.1f} games  (σ={SIGMA})")
            print(f"     Quota        : {v['quota_handicap']} | Prob: {v['prob_stimata']*100:.1f}%")
            print(f"     EV           : {v['ev']*100:.1f}%")
            print()
    else:
        print("  Nessuna value bet handicap trovata.")

    if senza_quote:
        print(f"\n⚠️  Senza quote handicap ({len(senza_quote)}):")
        for s in senza_quote[:5]:
            print(f"  - {s}")

    if non_trovati:
        print(f"\n⚠️  Non trovati in Elo ({len(non_trovati)}):")
        for n in non_trovati[:5]:
            print(f"  - {n}")

    salva_handicap_excel(value_bets)
    return value_bets


if __name__ == "__main__":
    risultati = scansiona_handicap(ev_max=EV_MAX)

    # ── Riepilogo finale compatto ────────────────────────────────────────
    print()
    print("=" * 65)
    print("  RIEPILOGO FINALE — SCANNER HANDICAP")
    print("=" * 65)
    print(f"  Data scansione  : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"  VALUE BET totali: {len(risultati)}")

    if risultati:
        print()
        print(f"  {'Giocatore (handicap)':<36} {'vs Avversario':<22} {'H':>5}  {'Quota':>6}  {'EV%':>6}  {'Marg':>7}")
        print(f"  {'-'*36} {'-'*22} {'-'*5}  {'-'*6}  {'-'*6}  {'-'*7}")
        for v in risultati:
            tipo = v.get('tipo', 'Games')
            nome_hc = f"{v['p1']} ({v['handicap']:+.1f} {tipo.lower()})"
            marg = f"{v['margine_atteso']*100:.1f}%" if tipo == 'Sets' else f"{v['margine_atteso']:>+4.1f}g"
            print(
                f"  {nome_hc:<36} {v['p2']:<22} "
                f"{v['handicap']:>+5.1f}  {v['quota_handicap']:>6.3f}  "
                f"{v['ev']*100:>5.1f}%  {marg}"
            )
    else:
        print("\n  Nessuna value bet handicap trovata oggi.")

    print("=" * 65)
