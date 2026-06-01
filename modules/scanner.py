import requests
import sys
import os
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from datetime import datetime
from modules.elo_tennisabstract import (
    carica_elo_aggiornato, trova_giocatore_ta, verifica_disponibilita
)
from modules.odds_apiio import get_partite_con_quote_oggi, get_quote_evento as get_quote_apiio
from modules.signals import genera_segnale, prob_da_elo
from modules.forma_recente import carica_partite_recenti, calcola_forma, aggiusta_elo_per_forma, calcola_h2h, calcola_fatica
from modules.data_2025 import scarica_e_salva_2026
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SOFASCORE_URL = "https://api.sofascore.com/api/v1/sport/tennis/scheduled-events/{data}"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

try:
    from config import (ODDS_API_IO_KEY, EV_MAX, EV_MINIMO, ODDS_MIN_VALUE, ODDS_MAX_VALUE,
                        MODELLO, MARKOV_CPI_ABILITATO, MARKOV_FATICA_ESTESA,
                        MARKOV_TIMEZONE_ABILITATO, MARKOV_ETA_SUPERFICIE, MARKOV_CPI)
except Exception:
    ODDS_API_IO_KEY = ""
    EV_MAX = None
    EV_MINIMO = 0.09
    ODDS_MIN_VALUE = 1.80
    ODDS_MAX_VALUE = 4.00
    MODELLO = "elo"
    MARKOV_CPI_ABILITATO = False
    MARKOV_FATICA_ESTESA = True
    MARKOV_TIMEZONE_ABILITATO = False
    MARKOV_ETA_SUPERFICIE = True
    MARKOV_CPI = {'hard': 0.0, 'clay': 0.0, 'grass': 0.0}

# Markov sempre importato: calcoliamo sempre entrambi i modelli (Elo + Markov)
# così l'Excel mostra le due quote in parallelo. MODELLO determina solo quale
# modello guida la soglia EV per il flag "value bet".
from modules.markov import calcola_probabilita_markov, calcola_fatica_markov

SUPERFICIE_MAP = {
    'clay': 'clay', 'red clay': 'clay', 'clay (red)': 'clay',
    'hard': 'hard', 'hardcourt outdoor': 'hard', 'hardcourt indoor': 'hard',
    'hard (indoor)': 'hard', 'grass': 'grass', 'carpet': 'hard',
}

def normalizza_superficie(s):
    if not s or str(s).lower() in ['nan','gwangju','']:
        return 'hard'
    return SUPERFICIE_MAP.get(str(s).strip().lower(), 'hard')

def superficie_da_torneo(torneo):
    t = torneo.lower()

    # GRASS — controllo che 'halle' non sia dentro 'challenger'
    if 'halle' in t and 'challenger' not in t:
        return 'grass'
    if 'queens club' in t:
        return 'grass'
    if any(x in t for x in [
        'wimbledon', 'hertogenbosch', 'eastbourne',
        'nottingham', 'newport', 'rosmalen'
    ]):
        return 'grass'

    # CLAY
    if any(x in t for x in [
        'madrid', 'rome', 'roma', 'roland', 'clay',
        'barcelona', 'monte carlo', 'hamburg', 'geneva',
        'lyon', 'munich', 'estoril', 'bucharest', 'istanbul',
        'marrakech', 'houston', 'rio', 'buenos aires', 'bogota',
        'santiago', 'lima', 'cordoba', 'bastad', 'umag', 'gstaad',
        'kitzbuhel', 'casablanca', 'cagliari', 'aix',
        'mauthausen', 'ostrava', 'abidjan', 'shymkent'
    ]):
        return 'clay'

    # HARD default
    return 'hard'

def get_quote_sofascore(event_id):
    url = f"https://api.sofascore.com/api/v1/event/{event_id}/odds/1/featured"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=8)
        data = resp.json()
        featured = data.get("featured", {}).get("default", {})
        choices = featured.get("choices", [])
        if len(choices) >= 2:
            def fraz_a_dec(fraz):
                try:
                    n, d = fraz.split("/")
                    return round(1 + int(n) / int(d), 3)
                except:
                    return None
            # Sofascore usa name="1" per home e name="2" per away.
            # Cercare per nome evita errori se l'ordine posizionale varia.
            home_c = next((c for c in choices if c.get("name") == "1"), None)
            away_c = next((c for c in choices if c.get("name") == "2"), None)
            if home_c and away_c:
                return (fraz_a_dec(home_c.get("fractionalValue", "")),
                        fraz_a_dec(away_c.get("fractionalValue", "")))
            # Fallback posizionale se i nomi non corrispondono allo schema atteso
            return fraz_a_dec(choices[0].get("fractionalValue", "")), fraz_a_dec(choices[1].get("fractionalValue", ""))
    except:
        pass
    return None, None

def get_partite_sofascore(data=None):
    if data is None:
        data = datetime.now().strftime("%Y-%m-%d")
    try:
        resp = requests.get(
            SOFASCORE_URL.format(data=data),
            headers=HEADERS, timeout=10
        )
        eventi = resp.json().get("events", [])
    except:
        return []

    partite = []
    for e in eventi:
        filtri = e.get("eventFilters", {})
        if "singles" not in filtri.get("category", []):
            continue
        if "pro" not in filtri.get("level", []):
            continue
        tornei_validi = ['p1000', 'p500', 'p250', 'grand_slam', 'atp', 'wta', 'challenger']
        if not any(t in filtri.get("tournament", []) for t in tornei_validi):
            continue
        if e.get("status", {}).get("type") == "finished":
            continue

        p1 = e.get("homeTeam", {}).get("name", "")
        p2 = e.get("awayTeam", {}).get("name", "")
        torneo = e.get("uniqueTournament", {}).get("name", "")
        superficie = e.get("groundType", "")
        event_id = e.get("id")

        if p1 and p2:
            partite.append({
                "id": event_id,
                "p1": p1, "p2": p2,
                "torneo": torneo,
                "superficie": superficie,
                "source": "sofascore"
            })
    return partite

def salva_value_bets_excel(value_bets, filepath="data/value_bets_log.xlsx"):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    # Nuovo header con doppia colonna modello (Elo + Markov)
    HEADERS = [
        "Data", "Ora", "🎯 PUNTA SU", "Avversario",
        "Torneo", "Superficie", "Quota",
        # Blocco Elo
        "Prob Elo %", "Quota Equa Elo", "EV Elo %",
        # Blocco Markov
        "Prob Markov %", "Quota Equa Markov", "EV Markov %",
        # Metadati e tracking
        "Elo Usato", "Fonte",
        "Quota Apertura", "Quota Chiusura", "CLV %",
        "Esito", "Profitto"
    ]

    verde_scuro  = PatternFill("solid", fgColor="1E7145")
    verde_chiaro = PatternFill("solid", fgColor="C6EFCE")
    bianco       = PatternFill("solid", fgColor="FFFFFF")
    giallo       = PatternFill("solid", fgColor="FFEB9C")
    azzurro      = PatternFill("solid", fgColor="BDD7EE")
    arancione_h  = PatternFill("solid", fgColor="ED7D31")
    verde_clv_h  = PatternFill("solid", fgColor="70AD47")
    blu_ap_h     = PatternFill("solid", fgColor="4472C4")
    arancione_d  = PatternFill("solid", fgColor="FCE4D6")
    verde_clv_d  = PatternFill("solid", fgColor="E2EFDA")
    # Header tinted per i due blocchi modello
    blu_elo_h    = PatternFill("solid", fgColor="305496")  # blu scuro = Elo
    blu_elo_d    = PatternFill("solid", fgColor="D9E1F2")  # blu chiaro dati
    viola_mk_h   = PatternFill("solid", fgColor="7030A0")  # viola scuro = Markov
    viola_mk_d   = PatternFill("solid", fgColor="E4D7F0")  # viola chiaro dati
    bordo = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _scrivi_header(ws):
        # colonne speciali: (fill, font_color)
        header_speciali = {
            8:  (blu_elo_h,    "FFFFFF"),  # Prob Elo
            9:  (blu_elo_h,    "FFFFFF"),  # Quota Equa Elo
            10: (blu_elo_h,    "FFFFFF"),  # EV Elo
            11: (viola_mk_h,   "FFFFFF"),  # Prob Markov
            12: (viola_mk_h,   "FFFFFF"),  # Quota Equa Markov
            13: (viola_mk_h,   "FFFFFF"),  # EV Markov
            16: (blu_ap_h,     "FFFFFF"),  # Quota Apertura
            17: (arancione_h,  "FFFFFF"),  # Quota Chiusura
            18: (verde_clv_h,  "FFFFFF"),  # CLV %
        }
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            fill_h, fc = header_speciali.get(col, (verde_scuro, "FFFFFF"))
            cell.fill = fill_h
            cell.font = Font(bold=True, color=fc, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = bordo
        ws.row_dimensions[1].height = 30

    # Verifica se il file esistente ha la vecchia struttura (senza colonne Markov)
    formato_vecchio = False
    if os.path.exists(filepath):
        try:
            wb_check = load_workbook(filepath)
            ws_check = wb_check.active
            headers_attuali = [ws_check.cell(1, c).value for c in range(1, ws_check.max_column + 1)]
            if "Prob Markov %" not in headers_attuali or "Quota Equa Elo" not in headers_attuali:
                formato_vecchio = True
        except Exception:
            formato_vecchio = True

    if formato_vecchio:
        # Backup del vecchio file e ripartenza con il nuovo schema
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = filepath.replace(".xlsx", f"_backup_pre_markov_{ts}.xlsx")
        try:
            os.rename(filepath, backup_path)
            print(f"  📦 Vecchio Excel salvato come backup: {backup_path}")
        except Exception as e:
            print(f"  ⚠️  Backup vecchio Excel fallito ({e}), procedo sovrascrivendo")
            try:
                os.remove(filepath)
            except Exception:
                pass

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active

        # Raccogli TUTTE le partite presenti (indipendentemente dalla data)
        # e rimuovi eventuali duplicati già nell'Excel
        partite_esistenti = set()
        righe_duplicate = []
        for row_idx in range(2, ws.max_row + 1):
            data_val = ws.cell(row_idx, 1).value
            p1_val   = ws.cell(row_idx, 3).value
            p2_val   = ws.cell(row_idx, 4).value
            if not data_val:
                continue
            data_str = data_val.strftime("%Y-%m-%d") if hasattr(data_val, 'strftime') else str(data_val).strip()
            raw_p1 = str(p1_val).replace('✅ ', '').strip() if p1_val else ''
            raw_p2 = str(p2_val).strip() if p2_val else ''
            key     = (data_str, raw_p1, raw_p2)
            key_inv = (data_str, raw_p2, raw_p1)
            if key in partite_esistenti or key_inv in partite_esistenti:
                righe_duplicate.append(row_idx)
            else:
                partite_esistenti.add(key)

        for row_idx in reversed(righe_duplicate):
            ws.delete_rows(row_idx)
        if righe_duplicate:
            print(f"  🧹 {len(righe_duplicate)} righe duplicate rimosse dall'Excel")

        prossima_riga = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Value Bets Log"
        partite_esistenti = set()
        _scrivi_header(ws)
        prossima_riga = 2

    ora_now = datetime.now()
    aggiunte = 0
    saltate = 0

    def _pct(x):
        try:
            return round(float(x) * 100, 1)
        except (TypeError, ValueError):
            return ""

    for i, v in enumerate(value_bets):
        giocatore = v['p1']
        avversario = v['p2']

        # Controlla duplicato su tutta la storia (non solo oggi)
        data_p = v.get('data_partita', '').strip()
        key = (data_p, giocatore.strip(), avversario.strip())
        key_inv = (data_p, avversario.strip(), giocatore.strip())
        if key in partite_esistenti or key_inv in partite_esistenti:
            saltate += 1
            continue

        partite_esistenti.add((data_p, giocatore.strip(), avversario.strip()))
        riga = prossima_riga + aggiunte
        fill = verde_chiaro if aggiunte % 2 == 0 else bianco

        # Fallback: se mancano i campi nuovi (vecchi dict), riusa prob_reale/ev/quota_equa
        prob_elo_v   = v.get('prob_elo',   v.get('prob_reale'))
        equa_elo_v   = v.get('quota_equa_elo',   v.get('quota_equa'))
        ev_elo_v     = v.get('ev_elo',     v.get('ev'))
        prob_mk_v    = v.get('prob_markov',     v.get('prob_reale'))
        equa_mk_v    = v.get('quota_equa_markov', v.get('quota_equa'))
        ev_mk_v      = v.get('ev_markov',  v.get('ev'))

        valori = [
            v.get('data_partita', ora_now.strftime("%Y-%m-%d")),
            ora_now.strftime("%H:%M"),
            f"✅ {giocatore}",          # Chi scommettere
            avversario,
            v['torneo'],
            v['superficie'].upper(),
            v['quota_p1'],              # Quota mercato attuale
            # ── Blocco Elo ──
            _pct(prob_elo_v),
            equa_elo_v,
            _pct(ev_elo_v),
            # ── Blocco Markov ──
            _pct(prob_mk_v),
            equa_mk_v,
            _pct(ev_mk_v),
            # ── Metadati / tracking ──
            v.get('elo_usato', ''),
            v.get('source', ''),
            v['quota_p1'],              # Quota Apertura (snapshot al momento scoperta)
            "",                         # Quota Chiusura (manuale)
            "",                         # CLV % (manuale)
            "",
            ""
        ]

        for col, val in enumerate(valori, 1):
            cell = ws.cell(row=riga, column=col, value=val)
            cell.border = bordo
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Colonna "PUNTA SU" in azzurro con font bold
            if col == 3:
                cell.fill = azzurro
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # Quota mercato
            elif col == 7:
                cell.fill = fill
                cell.font = Font(bold=True, size=12)
            # ── Blocco Elo (8=Prob, 9=QuotaEqua, 10=EV) ─────────────────
            elif col in (8, 9):
                cell.fill = blu_elo_d
            elif col == 10 and isinstance(val, (int, float)):
                # EV Elo colorata per intensità
                if val >= 50:
                    cell.fill = PatternFill("solid", fgColor="FF0000")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val >= 20:
                    cell.fill = PatternFill("solid", fgColor="FFC000")
                    cell.font = Font(bold=True)
                elif val >= 10:
                    cell.fill = giallo
                else:
                    cell.fill = blu_elo_d
            # ── Blocco Markov (11=Prob, 12=QuotaEqua, 13=EV) ────────────
            elif col in (11, 12):
                cell.fill = viola_mk_d
            elif col == 13 and isinstance(val, (int, float)):
                if val >= 50:
                    cell.fill = PatternFill("solid", fgColor="FF0000")
                    cell.font = Font(bold=True, color="FFFFFF")
                elif val >= 20:
                    cell.fill = PatternFill("solid", fgColor="FFC000")
                    cell.font = Font(bold=True)
                elif val >= 10:
                    cell.fill = giallo
                else:
                    cell.fill = viola_mk_d
            # Quota Apertura
            elif col == 16:
                cell.fill = fill
                cell.font = Font(bold=True, size=12)
            # Quota Chiusura: sfondo arancione chiaro
            elif col == 17:
                cell.fill = arancione_d
            # CLV %: sfondo verde chiaro
            elif col == 18:
                cell.fill = verde_clv_d
            else:
                cell.fill = fill

        ws.row_dimensions[riga].height = 22
        aggiunte += 1

    # Larghezze colonne (19 colonne totali)
    larghezze = [
        12, 8, 28, 22, 30, 10, 8,            # base
        10, 14, 10,                          # Elo
        12, 16, 12,                          # Markov
        14, 12, 14, 14, 10, 8, 10            # metadati + tracking
    ]
    for col, width in enumerate(larghezze, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    wb.save(filepath)
    print(f"\n💾 {aggiunte} nuove value bet salvate | {saltate} duplicati saltati → {filepath}")

def analizza_partite(partite, ratings_ta, soglia_ev, get_quote_fn, partite_recenti=None, nomi_recenti_dict=None):
    value_bets = []
    no_value = []
    non_trovati = []
    senza_quote = []

    # ── Rate limiter: max 10 richieste al secondo ──────────────────────
    _rl_lock = threading.Lock()
    _rl_slots = []

    def _rate_limit():
        with _rl_lock:
            while True:
                now = time.time()
                _rl_slots[:] = [t for t in _rl_slots if now - t < 1.0]
                if len(_rl_slots) < 10:
                    break
                time.sleep(1.0 - (now - _rl_slots[0]) + 0.001)
            _rl_slots.append(time.time())

    def _throttled_quote(event_id):
        _rate_limit()
        return get_quote_fn(event_id)

    # ── FASE 1: calcolo Elo per tutte le partite (sequenziale, in-memory) ──
    partite_valide = []
    for p in partite:
        sup_raw = p.get('superficie', '')
        sup = normalizza_superficie(sup_raw)
        if sup == 'hard' and not sup_raw:
            sup = superficie_da_torneo(p.get('torneo', ''))

        n1, r1 = trova_giocatore_ta(p['p1'], ratings_ta)
        n2, r2 = trova_giocatore_ta(p['p2'], ratings_ta)

        if not r1 or not r2:
            mancante = p['p1'] if not r1 else p['p2']
            non_trovati.append(f"{p['p1']} vs {p['p2']} (manca: {mancante})")
            continue

        e1 = r1.get(sup, r1['elo'])
        e2 = r2.get(sup, r2['elo'])
        elo_usato = f"TA-{sup}"

        n1_fr = None
        n2_fr = None
        if partite_recenti and nomi_recenti_dict:
            n1_fr, _ = trova_giocatore_ta(n1, nomi_recenti_dict)
            n2_fr, _ = trova_giocatore_ta(n2, nomi_recenti_dict)
            forma1 = calcola_forma(n1_fr, partite_recenti, superficie=sup) if n1_fr else 0.0
            forma2 = calcola_forma(n2_fr, partite_recenti, superficie=sup) if n2_fr else 0.0
            e1 = aggiusta_elo_per_forma(e1, forma1)
            e2 = aggiusta_elo_per_forma(e2, forma2)
            elo_usato = f"TA-{sup}+forma"

        h2h = calcola_h2h(n1, n2, partite_recenti)
        # H2H aggiusta Elo di max ±50 punti (ridotto da 75; pochi dati storici su Challenger)
        e1 = e1 + (h2h * 50)
        e2 = e2 - (h2h * 50)
        if h2h != 0:
            elo_usato += f"+h2h({h2h:+.2f})"

        if partite_recenti:
            fatica1 = calcola_fatica(n1_fr if n1_fr else n1, partite_recenti)
            fatica2 = calcola_fatica(n2_fr if n2_fr else n2, partite_recenti)
            # Fatica aggiusta Elo di max -100 punti; si attiva con 2+ partite negli ultimi 2 giorni
            e1 = e1 + (fatica1 * 100)
            e2 = e2 + (fatica2 * 100)
            if fatica1 != 0 or fatica2 != 0:
                elo_usato += f"+fat({fatica1:.1f}/{fatica2:.1f})"

        partite_valide.append({
            'p': p, 'n1': n1, 'n2': n2,
            'e1': e1, 'e2': e2,
            'sup': sup, 'elo_usato': elo_usato,
            'n1_fr': n1_fr, 'n2_fr': n2_fr,  # nomi matched su partite_recenti
            'eta1': r1.get('age', 0) or 0,
            'eta2': r2.get('age', 0) or 0,
        })

    # ── FASE 2: fetch quote in parallelo (max 10 worker, rate-limited) ──
    quote_cache = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {
            executor.submit(_throttled_quote, item['p']['id']): item['p']['id']
            for item in partite_valide
        }
        for future in as_completed(futures):
            eid = futures[future]
            try:
                quote_cache[eid] = future.result()
            except Exception:
                quote_cache[eid] = (None, None)

    # ── FASE 3: segnali, filtri range Pro, value bet ──────────────────
    for item in partite_valide:
        p = item['p']
        n1, n2 = item['n1'], item['n2']
        e1, e2 = item['e1'], item['e2']
        sup, elo_usato = item['sup'], item['elo_usato']
        n1_fr = item.get('n1_fr')
        n2_fr = item.get('n2_fr')

        q1, q2 = quote_cache.get(p['id'], (None, None))

        if not q1 or not q2:
            senza_quote.append(f"{p['p1']} vs {p['p2']}")
            continue

        # Guard: se le due quote sono quasi identiche la normalizzazione
        # home/away non è riuscita a risolvere l'ambiguità (split 50/50
        # tra bookmaker). Meglio saltare che rischiare un segnale falso.
        if abs(q1 - q2) < 0.15:
            print(f"  ⚠️  Quote ambigue scartate: {p['p1']} vs {p['p2']} ({q1} / {q2})")
            senza_quote.append(f"{p['p1']} vs {p['p2']} (quote ambigue: {q1}/{q2})")
            continue

        # Sanity check Elo: quota altissima ma l'Elo favorisce quel giocatore → quote invertite
        if q1 > 8.0 and e1 > e2:
            print(f"  🔄 Quote invertite per sanity check Elo: {p['p1']} vs {p['p2']} ({q1} → {q2})")
            q1, q2 = q2, q1
        elif q2 > 8.0 and e2 > e1:
            print(f"  🔄 Quote invertite per sanity check Elo: {p['p1']} vs {p['p2']} ({q1} → {q2})")
            q1, q2 = q2, q1

        # Filtro range Pro: scarta se entrambe le quote sono fuori range
        q1_ok = ODDS_MIN_VALUE <= q1 <= ODDS_MAX_VALUE
        q2_ok = ODDS_MIN_VALUE <= q2 <= ODDS_MAX_VALUE

        if not q1_ok and not q2_ok:
            senza_quote.append(f"{p['p1']} vs {p['p2']} (quota fuori range Pro)")
            continue

        trovata = False
        entry = None

        # ── Probabilità: SEMPRE entrambi i modelli (Elo + Markov) ────────
        # Elo: derivato direttamente dai rating Tennis Abstract (con adj forma/h2h/fatica)
        prob_elo_1 = prob_da_elo(e1, e2)
        prob_elo_2 = prob_da_elo(e2, e1)

        # Markov: con adjustment contestuali (CPI, fatica 14gg, età × superficie)
        adj_1: dict = {}
        adj_2: dict = {}
        if MARKOV_CPI_ABILITATO:
            adj_1['cpi'] = MARKOV_CPI.get(sup, 0.0)
            adj_2['cpi'] = MARKOV_CPI.get(sup, 0.0)
        if MARKOV_FATICA_ESTESA and partite_recenti:
            n1_key = n1_fr if n1_fr else n1
            n2_key = n2_fr if n2_fr else n2
            adj_1['fatica_A'] = calcola_fatica_markov(n1_key, partite_recenti, giorni=14)
            adj_1['fatica_B'] = calcola_fatica_markov(n2_key, partite_recenti, giorni=14)
            adj_2['fatica_A'] = adj_1['fatica_B']
            adj_2['fatica_B'] = adj_1['fatica_A']
        if MARKOV_ETA_SUPERFICIE:
            eta1 = item.get('eta1', 0)
            eta2 = item.get('eta2', 0)
            if eta1:
                adj_1['eta_A'] = eta1
                adj_2['eta_B'] = eta1
            if eta2:
                adj_1['eta_B'] = eta2
                adj_2['eta_A'] = eta2
        prob_markov_1 = calcola_probabilita_markov(e1, e2, sup, best_of=3, adjustments=adj_1)
        prob_markov_2 = calcola_probabilita_markov(e2, e1, sup, best_of=3, adjustments=adj_2)

        # Modello primario (guida la soglia EV per il flag value bet)
        if MODELLO == "markov":
            prob_1, prob_2 = prob_markov_1, prob_markov_2
            modello_tag = "markov"
        else:
            prob_1, prob_2 = prob_elo_1, prob_elo_2
            modello_tag = "elo"

        # Helper: dato prob + quota, ritorna (quota_equa, ev)
        def _equa_ev(prob: float, quota: float) -> tuple:
            equa = round(1.0 / prob, 2) if prob > 0 else 99
            ev = round((prob * quota) - 1.0, 4)
            return equa, ev

        # Calcola segnale P1 solo se q1 in range
        if q1_ok:
            segnale = genera_segnale(n1, e1, n2, e2, q1, q2, prob_override=prob_1)
            quota_equa_elo_1, ev_elo_1       = _equa_ev(prob_elo_1, q1)
            quota_equa_markov_1, ev_markov_1 = _equa_ev(prob_markov_1, q1)
            entry = {
                "p1": p['p1'], "p2": p['p2'],
                "torneo": p['torneo'],
                "superficie": sup,
                "quota_p1": q1, "quota_p2": q2,
                "prob_reale": segnale['prob_reale'],
                "ev": segnale['ev'],
                "quota_equa": segnale['quota_equa'],
                # Doppio modello (sempre presenti)
                "prob_elo": round(prob_elo_1, 4),
                "quota_equa_elo": quota_equa_elo_1,
                "ev_elo": ev_elo_1,
                "prob_markov": round(prob_markov_1, 4),
                "quota_equa_markov": quota_equa_markov_1,
                "ev_markov": ev_markov_1,
                "modello_primario": modello_tag,
                "elo_usato": f"{elo_usato} [{modello_tag}]",
                "source": p.get('source', ''),
                "data_partita": p.get("data_partita", ""),
            }
            if segnale['ev'] >= soglia_ev:
                value_bets.append(entry)
                trovata = True

        # Controlla P2 se P1 non è value (o q1 fuori range) e q2 in range
        if not trovata and q2_ok:
            segnale2 = genera_segnale(n2, e2, n1, e1, q2, q1, prob_override=prob_2)
            quota_equa_elo_2, ev_elo_2       = _equa_ev(prob_elo_2, q2)
            quota_equa_markov_2, ev_markov_2 = _equa_ev(prob_markov_2, q2)
            if segnale2['ev'] >= soglia_ev:
                value_bets.append({
                    "p1": p['p2'], "p2": p['p1'],
                    "torneo": p['torneo'],
                    "superficie": sup,
                    "quota_p1": q2, "quota_p2": q1,
                    "prob_reale": segnale2['prob_reale'],
                    "ev": segnale2['ev'],
                    "quota_equa": segnale2['quota_equa'],
                    "prob_elo": round(prob_elo_2, 4),
                    "quota_equa_elo": quota_equa_elo_2,
                    "ev_elo": ev_elo_2,
                    "prob_markov": round(prob_markov_2, 4),
                    "quota_equa_markov": quota_equa_markov_2,
                    "ev_markov": ev_markov_2,
                    "modello_primario": modello_tag,
                    "elo_usato": f"{elo_usato} [{modello_tag}]",
                    "source": p.get('source', ''),
                    "data_partita": p.get("data_partita", ""),
                })
            elif entry is not None:
                no_value.append(entry)
        elif not trovata and entry is not None:
            no_value.append(entry)

    return value_bets, no_value, non_trovati, senza_quote

def aggiorna_esiti_excel(
    csv_path='data/storico/risultati_recenti.csv',
    xlsx_path='data/value_bets_log.xlsx'
):
    if not os.path.exists(csv_path):
        print("  risultati_recenti.csv non trovato — skip aggiornamento esiti")
        return
    if not os.path.exists(xlsx_path):
        print("  value_bets_log.xlsx non trovato — skip aggiornamento esiti")
        return

    import pandas as pd
    df = pd.read_csv(csv_path)

    # Per ogni data costruisci set distinti di vincitori e perdenti
    winners_per_data = {}
    losers_per_data  = {}
    for _, row in df.iterrows():
        d = int(row['tourney_date'])
        winners_per_data.setdefault(d, set()).add(row['winner_name'])
        losers_per_data.setdefault(d, set()).add(row['loser_name'])

    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Rileva posizioni colonne dalla riga header
    header_map = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    col_data     = header_map.get("Data",         1)
    col_gioc     = header_map.get("🎯 PUNTA SU",  3)
    col_quota    = header_map.get("Quota",        7)
    col_esito    = header_map.get("Esito",        19)
    col_profitto = header_map.get("Profitto",     20)

    aggiornati = 0
    n_w = 0
    n_l = 0

    for row_idx in range(2, ws.max_row + 1):
        esito_val = ws.cell(row_idx, col_esito).value
        if esito_val not in (None, ''):
            continue

        data_val = ws.cell(row_idx, col_data).value
        gioc_val = ws.cell(row_idx, col_gioc).value
        if not data_val or not gioc_val:
            continue

        try:
            if isinstance(data_val, str):
                data_int = int(data_val.replace('-', ''))
            else:
                data_int = int(data_val.strftime('%Y%m%d'))
        except Exception:
            continue

        giocatore = str(gioc_val).replace('✅ ', '').strip()

        winners = winners_per_data.get(data_int, set())
        losers  = losers_per_data.get(data_int, set())
        tutti   = winners | losers
        if not tutti:
            continue

        nomi_dict = {n: {} for n in tutti}
        nome_match, _ = trova_giocatore_ta(giocatore, nomi_dict)
        if nome_match is None:
            continue

        if nome_match in winners:
            esito = 'W'
            try:
                quota = float(ws.cell(row_idx, col_quota).value)
                profitto = round(quota - 1, 2)
            except (TypeError, ValueError):
                profitto = 1
            n_w += 1
        else:
            esito, profitto = 'L', -1
            n_l += 1

        ws.cell(row_idx, col_esito).value    = esito
        ws.cell(row_idx, col_profitto).value = profitto
        aggiornati += 1

    wb.save(xlsx_path)
    print(f"\n📊 Esiti aggiornati: {aggiornati} ({n_w} W, {n_l} L) → {xlsx_path}")


def scansiona(soglia_ev=EV_MINIMO, ev_max=None):
    print(f"{'='*65}")
    print(f"  🎾 TENNIS SCANNER - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'='*65}\n")

    # ==========================================
    # SAFETY CHECK — blocca se fonti non disponibili
    # ==========================================
    print("🔒 Safety check fonti dati...")

    ok_ta, msg_ta = verifica_disponibilita()
    print(f"  Tennis Abstract : {'✅' if ok_ta else '❌'} {msg_ta}")

    ok_odds = False
    if ODDS_API_IO_KEY:
        try:
            r = requests.get(
                "https://api.odds-api.io/v3/sports",
                params={"apiKey": ODDS_API_IO_KEY}, timeout=5
            )
            ok_odds = r.status_code == 200
            print(f"  Odds-API.io     : {'✅' if ok_odds else '❌'} Status {r.status_code}")
        except Exception as e:
            print(f"  Odds-API.io     : ❌ {e}")

    if not ok_ta:
        print("\n🚨 BLOCCO: Tennis Abstract non disponibile.")
        print("   I rating Elo non sono aggiornati — scanner fermato.")
        print("   ⚠️  NON INVESTIRE senza dati Elo verificati.\n")
        return

    if not ok_odds:
        print("\n🚨 BLOCCO: Odds-API.io non disponibile.")
        print("   Le quote non sono recuperabili — scanner fermato.")
        print("   ⚠️  NON INVESTIRE senza quote verificate.\n")
        return

    print("\n✅ Tutte le fonti disponibili — procedo\n")

    # ==========================================
    # CARICA ELO AGGIORNATO
    # ==========================================
    print("⚙️  Caricamento Elo Tennis Abstract...")
    ratings_ta = carica_elo_aggiornato()
    if not ratings_ta:
        print("🚨 BLOCCO: Nessun dato Elo caricato — scanner fermato.")
        return
    print(f"✅ {len(ratings_ta)} giocatori caricati\n")

    # ==========================================
    # DATI 2026 + FORMA RECENTE
    # ==========================================
    csv_2026 = 'data/storico/atp_2026_tml.csv'
    deve_scaricare = (
        not os.path.exists(csv_2026)
        or (datetime.now() - datetime.fromtimestamp(os.path.getmtime(csv_2026))).total_seconds() > 86400
    )
    if deve_scaricare:
        print("📥 Aggiornamento dati 2026 da TML-Database...")
        try:
            scarica_e_salva_2026()
        except Exception as e:
            print(f"  ⚠️  Download 2026 fallito: {e}")
    else:
        print("📦 Dati 2026 già aggiornati (meno di 24h)")

    csv_recenti = 'data/storico/risultati_recenti.csv'
    deve_aggiornare = (
        not os.path.exists(csv_recenti)
        or (datetime.now() - datetime.fromtimestamp(
            os.path.getmtime(csv_recenti))).total_seconds() > 86400
    )
    if deve_aggiornare:
        print("Aggiornamento risultati recenti da Tennisexplorer...")
        try:
            from modules.data_2025 import scarica_risultati_recenti_tennisexplorer
            scarica_risultati_recenti_tennisexplorer(giorni=60)
        except Exception as e:
            print(f"  Download risultati recenti fallito: {e}")
    else:
        print("Risultati recenti gia aggiornati (meno di 24h)")

    print("📈 Caricamento forma recente (ultimi 30 giorni)...")
    partite_recenti = carica_partite_recenti(giorni=30)
    nomi_recenti_dict = {nome: {'elo': 0} for nome in partite_recenti}
    print(f"✅ Forma disponibile per {len(partite_recenti)} giocatori\n")

    # ==========================================
    # RECUPERA PARTITE
    # ==========================================
    all_value_bets = []
    all_no_value = []
    all_non_trovati = []
    all_senza_quote = []

    # Fonte 1: odds-api.io
    print("📅 Recupero partite da Odds-API.io...")
    partite_apiio = get_partite_con_quote_oggi()
    TORNEI_VALIDI = ['ATP -', 'WTA -', 'Challenger -']
    TORNEI_ESCLUDI = ['125K', '125k', 'UTR', 'ITF', 'WTA 125']

    partite_apiio = [
        p for p in partite_apiio
        if any(x in p['torneo'] for x in TORNEI_VALIDI)
        and not any(x in p['torneo'] for x in TORNEI_ESCLUDI)
    ]
    print(f"✅ {len(partite_apiio)} partite ATP/WTA trovate\n")

    vb, nv, nt, sq = analizza_partite(
        partite_apiio, ratings_ta, soglia_ev, get_quote_apiio,
        partite_recenti, nomi_recenti_dict
    )
    all_value_bets.extend(vb)
    all_no_value.extend(nv)
    all_non_trovati.extend(nt)
    all_senza_quote.extend(sq)

    # Fonte 2: Sofascore (per superficie e partite extra)
    print("📅 Recupero partite da Sofascore...")
    partite_sofa = get_partite_sofascore()
    print(f"✅ {len(partite_sofa)} partite trovate\n")

    nomi_apiio = {(p['p1'], p['p2']) for p in partite_apiio}
    partite_extra = [
        p for p in partite_sofa
        if (p['p1'], p['p2']) not in nomi_apiio
        and (p['p2'], p['p1']) not in nomi_apiio
    ]

    if partite_extra:
        print(f"✅ {len(partite_extra)} partite extra solo su Sofascore\n")
        vb2, nv2, nt2, sq2 = analizza_partite(
            partite_extra, ratings_ta, soglia_ev, get_quote_sofascore,
            partite_recenti, nomi_recenti_dict
        )
        all_value_bets.extend(vb2)
        all_no_value.extend(nv2)
        all_non_trovati.extend(nt2)
        all_senza_quote.extend(sq2)

    # Deduplicazione e ordinamento
    seen = set()
    value_bets_dedup = []
    for v in sorted(all_value_bets, key=lambda x: x['ev'], reverse=True):
        key = tuple(sorted([v['p1'], v['p2']]))
        if key not in seen:
            seen.add(key)
            value_bets_dedup.append(v)

    # Filtro EV massimo
    if ev_max is not None:
        prima = len(value_bets_dedup)
        value_bets_dedup = [v for v in value_bets_dedup if v['ev'] <= ev_max]
        escluse = prima - len(value_bets_dedup)
        if escluse:
            print(f"⚠️  {escluse} value bet escluse per EV > {ev_max*100:.0f}%\n")

    # ==========================================
    # OUTPUT
    # ==========================================
    print(f"{'='*65}")
    print(f"  RISULTATI SCANNER")
    print(f"{'='*65}")
    print(f"  Analizzate con quote : {len(value_bets_dedup) + len(all_no_value)}")
    print(f"  Senza quote          : {len(all_senza_quote)}")
    print(f"  Non trovati in Elo   : {len(all_non_trovati)}")
    print(f"  VALUE BET ✅         : {len(value_bets_dedup)}")
    print(f"{'='*65}\n")

    if value_bets_dedup:
        print("🎯 VALUE BET TROVATE:\n")
        for v in value_bets_dedup:
            print(f"  ✅ {v['p1']} vs {v['p2']}")
            print(f"     Torneo    : {v['torneo']}")
            print(f"     Superficie: {v['superficie']} | Elo: {v['elo_usato']}")
            print(f"     Quota     : {v['quota_p1']}")
            p_elo = v.get('prob_elo', v.get('prob_reale', 0)) or 0
            p_mk  = v.get('prob_markov', v.get('prob_reale', 0)) or 0
            q_elo = v.get('quota_equa_elo', v.get('quota_equa', ''))
            q_mk  = v.get('quota_equa_markov', v.get('quota_equa', ''))
            ev_elo = v.get('ev_elo', v.get('ev', 0)) or 0
            ev_mk  = v.get('ev_markov', v.get('ev', 0)) or 0
            print(f"     Elo    : prob {p_elo*100:.1f}% | equa {q_elo} | EV {ev_elo*100:+.1f}%")
            print(f"     Markov : prob {p_mk*100:.1f}% | equa {q_mk} | EV {ev_mk*100:+.1f}%")
            print()
    else:
        print("  Nessuna value bet trovata oggi.")

    if all_senza_quote:
        print(f"\n⚠️  Senza quote ({len(all_senza_quote)}):")
        for s in all_senza_quote[:5]:
            print(f"  - {s}")

    if all_non_trovati:
        print(f"\n⚠️  Non trovati in Elo ({len(all_non_trovati)}):")
        for n in all_non_trovati[:5]:
            print(f"  - {n}")

    salva_value_bets_excel(value_bets_dedup)
    aggiorna_esiti_excel()
    return value_bets_dedup

if __name__ == "__main__":
    scansiona(ev_max=EV_MAX)